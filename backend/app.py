"""
Natiqi Backend with SQLite Database
Complete version with all tables
"""
from flask import Flask, request, jsonify, Response
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import inspect, text, or_
from decimal import Decimal
import json
import io
import csv

from models import db, Admin, Specialist, Patient, RegisteredUser, EEGSession, EEGSessionEvent, Alert, SystemLog, Model, PatientSettings, Notification, NotificationEvent
import os
import sys
import uuid
import secrets
from datetime import date, datetime, timedelta
from pathlib import Path

# Windows consoles often use cp1252; set UTF-8 before any prints that may use non-ASCII
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

print(f'Python executable: {sys.executable}')

# ML inference (EEG 5-word SVM)
try:
    from ml.eeg_svm_5word import ModelNotReadyError, predict_window
except Exception as e:
    print(f'[WARN] ML import disabled: {e}')
    ModelNotReadyError = RuntimeError  # type: ignore
    predict_window = None  # type: ignore

# ML live demo (RF 4-word from LiveDataModels)
try:
    from ml.eeg_rf_4word_demo import ModelNotReadyError as LiveDemoModelNotReadyError, predict_live_demo
except Exception as e:
    print(f'[WARN] Live-demo ML import disabled: {e}')
    LiveDemoModelNotReadyError = RuntimeError  # type: ignore
    predict_live_demo = None  # type: ignore

# ══════════════════════════════════════════════════════════
# APP SETUP
# ══════════════════════════════════════════════════════════

app = Flask(__name__)

# SQLite database file location

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{os.path.join(BASE_DIR, 'natiqi.db')}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'natiqi-dev-secret'


# Local dev: Expo picks an ephemeral port (scripts/run-expo-web.cjs); allow any browser origin.
# For a locked-down production API, replace with an explicit origin list.
CORS(app, resources={r"/*": {"origins": "*"}})


# Initialize database
db.init_app(app)


def ensure_registered_user_verification_columns():
    """Add verification columns when upgrading an existing SQLite DB (db.create_all does not alter tables)."""
    try:
        inspector = inspect(db.engine)
        cols = {c['name'] for c in inspector.get_columns('registered_user')}
    except Exception:
        return
    with db.engine.begin() as conn:
        if 'is_verified' not in cols:
            conn.execute(text(
                'ALTER TABLE registered_user ADD COLUMN is_verified BOOLEAN NOT NULL DEFAULT 1'
            ))
        if 'verification_code' not in cols:
            conn.execute(text(
                'ALTER TABLE registered_user ADD COLUMN verification_code VARCHAR(10)'
            ))
        if 'verification_expires' not in cols:
            conn.execute(text(
                'ALTER TABLE registered_user ADD COLUMN verification_expires DATETIME'
            ))


def ensure_admin_specialist_gender_columns():
    """Add gender column when upgrading an existing SQLite DB."""
    for table in ('admin', 'specialist'):
        try:
            inspector = inspect(db.engine)
            cols = {c['name'] for c in inspector.get_columns(table)}
        except Exception:
            continue
        if 'gender' not in cols:
            with db.engine.begin() as conn:
                conn.execute(text(f'ALTER TABLE {table} ADD COLUMN gender VARCHAR(10)'))


def ensure_admin_specialist_verification_columns():
    """Add verification columns to admin/specialist when upgrading an older SQLite DB."""
    for table in ('admin', 'specialist'):
        try:
            inspector = inspect(db.engine)
            cols = {c['name'] for c in inspector.get_columns(table)}
        except Exception:
            continue
        alters: list[str] = []
        if 'verification_code' not in cols:
            alters.append(f'ALTER TABLE {table} ADD COLUMN verification_code VARCHAR(10)')
        if 'verification_expires' not in cols:
            alters.append(f'ALTER TABLE {table} ADD COLUMN verification_expires DATETIME')
        if not alters:
            continue
        with db.engine.begin() as conn:
            for stmt in alters:
                conn.execute(text(stmt))


def ensure_patient_room_numbers():
    """Assign unique random room numbers (4 digits) to patients with missing room (upgrade path)."""
    try:
        missing = Patient.query.filter(
            or_(Patient.room_number.is_(None), Patient.room_number == '')
        ).all()
    except Exception:
        return
    if not missing:
        return
    used = {
        str(r).strip()
        for (r,) in db.session.query(Patient.room_number).all()
        if r is not None and str(r).strip() != ''
    }
    try:
        for p in missing:
            for _ in range(500):
                cand = str(secrets.randbelow(9000) + 1000)
                if cand not in used:
                    p.room_number = cand
                    used.add(cand)
                    break
            else:
                cand = f'R{secrets.randbelow(90000) + 10000}'
                while cand in used:
                    cand = f'R{secrets.randbelow(90000) + 10000}'
                p.room_number = cand[:10]
                used.add(p.room_number)
        db.session.commit()
        print(f'🏠 Assigned room numbers to {len(missing)} patient(s)')
    except Exception as e:
        db.session.rollback()
        print(f'⚠️ ensure_patient_room_numbers: {e}')


def ensure_patient_password_column(default_password: str = 'user@123'):
    """Add patient.password when upgrading an existing SQLite DB, and backfill missing passwords."""
    try:
        inspector = inspect(db.engine)
        cols = {c['name'] for c in inspector.get_columns('patient')}
    except Exception:
        return

    with db.engine.begin() as conn:
        if 'password' not in cols:
            conn.execute(text('ALTER TABLE patient ADD COLUMN password VARCHAR(255)'))

    try:
        missing = Patient.query.filter(
            or_(Patient.password.is_(None), Patient.password == '')
        ).all()
    except Exception:
        return

    if not missing:
        return

    try:
        hashed = generate_password_hash(default_password)
        for p in missing:
            p.password = hashed
        db.session.commit()
        print(f'🔐 Set default password for {len(missing)} patient(s)')
    except Exception as e:
        db.session.rollback()
        print(f'⚠️ ensure_patient_password_column: {e}')


def ensure_patient_settings_recorded_data_usage_column():
    """Add recorded_data_usage_allowed when upgrading an existing SQLite DB."""
    try:
        inspector = inspect(db.engine)
        cols = {c['name'] for c in inspector.get_columns('patient_settings')}
    except Exception:
        return
    if 'recorded_data_usage_allowed' not in cols:
        with db.engine.begin() as conn:
            conn.execute(
                text(
                    'ALTER TABLE patient_settings ADD COLUMN recorded_data_usage_allowed '
                    'BOOLEAN NOT NULL DEFAULT 0'
                )
            )


def ensure_specialist_3030303030_demo_patient_and_sessions():
    """
    Ensure specialist Omar (3030303030) has patient 1616161616 (password 123) and recent EEG sessions
    so the specialist dashboard Patients tab and Recent Sessions list populate after upgrades.
    (Patient and RegisteredUser are separate tables; the same national ID may exist in both for demo.)
    """
    OMAR_NID = '3030303030'
    PATIENT_NID_INT = 1616161616
    PATIENT_NID_STR = str(PATIENT_NID_INT)

    try:
        spec = Specialist.query.filter(
            or_(Specialist.national_id == OMAR_NID, Specialist.national_id == int(OMAR_NID))
        ).first()
        if not spec:
            return

        spec_key = str(spec.national_id)

        # Keep Omar's main seed cohort assigned to him (ids may be stored as str or int in SQLite)
        for nid in (6060606060, 7070707070, 9292929292):
            p = Patient.query.filter(
                or_(Patient.national_id == str(nid), Patient.national_id == nid)
            ).first()
            if p:
                p.specialist_national_id = spec_key

        patient = Patient.query.filter(
            or_(Patient.national_id == PATIENT_NID_STR, Patient.national_id == PATIENT_NID_INT)
        ).first()

        if not patient:
            patient = Patient(
                national_id=PATIENT_NID_INT,
                name='Ayla Al-Naimi',
                role='patient',
                password=generate_password_hash('123'),
                date_of_birth=date(2000, 1, 16),
                gender='Female',
                room_number='1616',
                device='EPOC X',
                status='Active',
                specialist_national_id=spec_key,
            )
            db.session.add(patient)
            db.session.flush()
        else:
            patient.specialist_national_id = spec_key
            patient.password = generate_password_hash('123')
            if not (patient.room_number or '').strip():
                patient.room_number = '1616'

        model = Model.query.order_by(Model.model_id.asc()).first()
        if not model:
            db.session.commit()
            print('⚠️ ensure_specialist_3030303030_demo_patient: no Model row; skipped EEG sessions')
            return

        now = datetime.now()
        pnid = patient.national_id
        existing = EEGSession.query.filter(
            EEGSession.specialist_national_id == spec_key,
            or_(EEGSession.patient_national_id == pnid, EEGSession.patient_national_id == str(pnid)),
        ).count()

        if existing < 1:
            s1 = EEGSession(
                patient_national_id=pnid,
                specialist_national_id=spec_key,
                model_id=model.model_id,
                start_time=now - timedelta(hours=2),
                end_time=now - timedelta(hours=1, minutes=30),
                detected_word='Thirst',
                confidence_level=0.8900,
                device='EPOC X',
                channels=14,
                session_status='Ended',
            )
            s2 = EEGSession(
                patient_national_id=pnid,
                specialist_national_id=spec_key,
                model_id=model.model_id,
                start_time=now - timedelta(minutes=50),
                end_time=now - timedelta(minutes=20),
                detected_word='Medicine',
                confidence_level=0.8700,
                device='EPOC X',
                channels=14,
                session_status='Ended',
            )
            db.session.add_all([s1, s2])
            db.session.flush()
            db.session.add_all(
                [
                    EEGSessionEvent(session_id=s1.session_id, event_time=now - timedelta(hours=1, minutes=55), detected_word='Thirst', confidence=0.88),
                    EEGSessionEvent(session_id=s1.session_id, event_time=now - timedelta(hours=1, minutes=50), detected_word='Thirst', confidence=0.90),
                    EEGSessionEvent(session_id=s1.session_id, event_time=now - timedelta(hours=1, minutes=45), detected_word='Hunger', confidence=0.72),
                    EEGSessionEvent(session_id=s2.session_id, event_time=now - timedelta(minutes=45), detected_word='Medicine', confidence=0.87),
                    EEGSessionEvent(session_id=s2.session_id, event_time=now - timedelta(minutes=40), detected_word='Medicine', confidence=0.86),
                ]
            )

        db.session.commit()
        print('✅ ensure_specialist_3030303030_demo_patient: patient 1616161616 + sessions OK')
    except Exception as e:
        db.session.rollback()
        print(f'⚠️ ensure_specialist_3030303030_demo_patient: {e}')


def ensure_demo_patient_notifications_for_specialist():
    """Seed in-app notifications for patients assigned to Omar (3030303030) for demo dashboards and bell."""
    OMAR = '3030303030'
    try:
        spec = Specialist.query.filter(
            or_(Specialist.national_id == OMAR, Specialist.national_id == int(OMAR))
        ).first()
        if not spec:
            return
        spec_key = str(spec.national_id)
        patients = Patient.query.filter(
            or_(Patient.specialist_national_id == spec_key, Patient.specialist_national_id == int(spec_key))
        ).all()
        if not patients:
            return
        now = datetime.now()
        demos = [
            ('جوع', 0.86, False),
            ('عطش', 0.88, False),
            ('دواء', 0.84, True),
            ('حمام', 0.81, False),
        ]
        added = 0
        for p in patients:
            pid = str(p.national_id)
            cnt = Notification.query.filter(
                or_(Notification.patient_national_id == pid, Notification.patient_national_id == p.national_id)
            ).count()
            if cnt >= 4:
                continue
            need = 4 - int(cnt)
            for i in range(need):
                word, conf, seen = demos[i % len(demos)]
                db.session.add(
                    Notification(
                        patient_national_id=pid,
                        detected_word=word,
                        confidence=Decimal(str(conf)),
                        event_time=now - timedelta(minutes=45 + i * 11 + (hash(pid) % 7)),
                        seen=seen,
                    )
                )
                added += 1
        if added:
            db.session.commit()
            print(f'✅ ensure_demo_patient_notifications: added {added} notification row(s)')
        else:
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f'⚠️ ensure_demo_patient_notifications: {e}')


def generate_six_digit_code() -> str:
    return ''.join(secrets.choice('0123456789') for _ in range(6))


def write_system_log(user_name, user_national_id, event, role='unknown'):
    # Generate readable log ID: SYS-YYYYMMDD-XXXX
    today    = date.today().strftime('%Y%m%d')
    # Count how many logs exist today to get the sequence number
    count    = SystemLog.query.filter(
        SystemLog.log_id.like(f'SYS-{today}-%')
    ).count()
    log_id   = f'SYS-{today}-{str(count + 1).zfill(4)}'

    log = SystemLog(
        log_id           = log_id,
        user_name        = user_name,
        user_national_id = user_national_id,
        event            = event,
        role             = role,
    )
    db.session.add(log)
    db.session.commit()

# ══════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════

##======    Login   ======##

@app.route('/auth/login', methods=['POST'])

def login():

    """
    Login with database validation
    Checks Admin, Specialist, and RegisteredUser tables based on role
    """

    data = request.get_json()
    national_id = data.get('national_id', '')
    password    = data.get('password', '')
    role        = data.get('role', '')
    #role_norm   = str(role or '').strip()

    print(f"\n📥 Login attempt:")
    print(f"   ID: {national_id}")
    print(f"   Role: {role}")

    user = None

    # Find user based on role
    if role == 'admin':
        user = Admin.query.filter_by(national_id=national_id).first()
    elif role == 'specialist':
        user = Specialist.query.filter_by(national_id=national_id).first()
    elif role == 'RegisteredUser':
        user = RegisteredUser.query.filter_by(national_id=national_id).first() 
           
    if not user:
        print(f"❌ No {role} found with ID: {national_id}")
        return jsonify({'error': 'Invalid ID or password'}), 401
    
    # Check password
    if not check_password_hash(user.password, password):
        display = getattr(user, 'name', None) or national_id
        print(f"❌ Wrong password for user: {display}")
        return jsonify({'error': 'Invalid ID or password'}), 401

    if isinstance(user, RegisteredUser) and not user.is_verified:
        return jsonify({'error': 'Please verify your account with the code sent to your email (check dev hint on verify screen if testing).'}), 403
    
    print(f"✅ Login successful: {getattr(user, 'name', None) or national_id}")
    return jsonify(user.to_dict()), 200

##======    SignUp   ======##
@app.route('/auth/register', methods=['POST'])
def register():
    data = request.get_json()

    # 1. Extract fields from what TypeScript sent
    national_id = data.get('national_id')
    name = (data.get('name') or '').strip()
    phone_num = data.get('phone_num')
    email = data.get('email')
    password = data.get('password')
    gender = data.get('gender')

    # 2. Check required fields
    if not all([national_id, name, email, password, phone_num, gender]):
        return jsonify({'error': 'All fields are required'}), 400
    if gender not in ('Male', 'Female'):
        return jsonify({'error': 'Gender must be Male or Female'}), 400

    existing = RegisteredUser.query.filter_by(national_id=national_id).first()
    # Verified accounts cannot be replaced
    if existing and existing.is_verified:
        return jsonify({'error': 'National ID already registered'}), 409

    # Email / phone must stay unique across other rows (allow same national_id to fix typos)
    email_owner = RegisteredUser.query.filter_by(email=email).first()
    if email_owner and email_owner.national_id != national_id:
        return jsonify({'error': 'Email already registered'}), 409
    phone_owner = RegisteredUser.query.filter_by(phone_num=phone_num).first()
    if phone_owner and phone_owner.national_id != national_id:
        return jsonify({'error': 'phone_num already registered'}), 409

    code = generate_six_digit_code()
    expires = datetime.utcnow() + timedelta(minutes=15)
    pw_hash = generate_password_hash(password)

    if existing:
        # Same ID, not verified yet — update details and issue a fresh code
        existing.name = name
        existing.phone_num = phone_num
        existing.email = email
        existing.password = pw_hash
        existing.gender = gender
        existing.verification_code = code
        existing.verification_expires = expires
        db.session.commit()
        msg = 'Details updated. Verify with the new code sent to your email (or dev hint below).'
    else:
        new_user = RegisteredUser(
            national_id=national_id,
            name=name,
            phone_num=phone_num,
            email=email,
            password=pw_hash,
            gender=gender,
            is_verified=False,
            verification_code=code,
            verification_expires=expires,
        )
        db.session.add(new_user)
        db.session.commit()
        msg = 'Account created successfully!'

    print(f"\n📧 Verification code for {email} (national_id={national_id}): {code}\n")

    payload = {'message': msg}
    # Local dev only: real app would email/SMS this code; never enable in production as-is.
    if app.debug or os.environ.get('RETURN_VERIFICATION_CODE') == '1':
        payload['dev_code'] = code
    return jsonify(payload), (200 if existing else 201)

##======    Verification   ======##

@app.route('/auth/verify', methods=['POST'])
def verify_account():
    data = request.get_json() or {}
    national_id = str(data.get('national_id', '')).strip()
    code = str(data.get('code', '')).strip()
    if not national_id or not code:
        return jsonify({'error': 'National ID and code are required'}), 400
    user = RegisteredUser.query.filter_by(national_id=national_id).first()
    if not user:
        return jsonify({'error': 'User not found'}), 404
    if user.is_verified:
        return jsonify({'message': 'Account is already verified'}), 200
    if not user.verification_code or user.verification_code != code:
        return jsonify({'error': 'Invalid verification code'}), 400
    if user.verification_expires and datetime.utcnow() > user.verification_expires:
        return jsonify({'error': 'Verification code has expired. Request a new code.'}), 400
    user.is_verified = True
    user.verification_code = None
    user.verification_expires = None
    db.session.commit()
    return jsonify({'message': 'Account verified. You can log in.'}), 200


##======    resend-verification   ======##

@app.route('/auth/resend-verification', methods=['POST'])
def resend_verification():
    data = request.get_json() or {}
    national_id = str(data.get('national_id', '')).strip()
    if not national_id:
        return jsonify({'error': 'National ID is required'}), 400
    user = RegisteredUser.query.filter_by(national_id=national_id).first()
    if not user:
        return jsonify({'error': 'User not found'}), 404
    if user.is_verified:
        return jsonify({'error': 'Account is already verified'}), 400
    code = generate_six_digit_code()
    user.verification_code = code
    user.verification_expires = datetime.utcnow() + timedelta(minutes=15)
    db.session.commit()
    print(f"\n📧 New verification code for {user.email} (national_id={national_id}): {code}\n")
    payload = {'message': 'A new verification code was generated.'}
    if app.debug or os.environ.get('RETURN_VERIFICATION_CODE') == '1':
        payload['dev_code'] = code
    return jsonify(payload), 200

##======    Forget Password   ======##
@app.route('/auth/forgot-password', methods=['POST'])
def forgot_password():
    """Step 1: User provides their national_id + role → generate a reset code."""
    data        = request.get_json() or {}
    national_id = str(data.get('national_id', '') or '').strip()
    role        = str(data.get('role', '') or '').strip()

    if not national_id or not role:
        return jsonify({'error': 'national_id and role are required'}), 400

    if role == 'admin':
        user = Admin.query.filter_by(national_id=national_id).first()
    elif role == 'specialist':
        user = Specialist.query.filter_by(national_id=national_id).first()
    elif role == 'RegisteredUser':
        user = RegisteredUser.query.filter_by(national_id=national_id).first()
    else:
        return jsonify({'error': 'Invalid role'}), 400

    if not user:
        # Don't reveal whether the ID exists
        return jsonify({'error': 'No account found with this ID.'}), 404
    
    code    = generate_six_digit_code()
    expires = datetime.utcnow() + timedelta(minutes=15)

    user.verification_code    = code
    user.verification_expires = expires
    db.session.commit()

    print(f"\n🔑 Password reset code for national_id={national_id}: {code}\n")

    payload = {'message': 'If this ID is registered, a reset code has been generated.'}
    if app.debug or os.environ.get('RETURN_VERIFICATION_CODE') == '1':
        payload['dev_code'] = code
    return jsonify(payload), 200


##======    Reset Password   ======##

@app.route('/auth/reset-password', methods=['POST'])
def reset_password():
    """Step 2: User provides national_id + role + code + new_password → update password."""
    data         = request.get_json() or {}
    national_id  = str(data.get('national_id', '') or '').strip()
    role         = str(data.get('role', '') or '').strip()
    code         = str(data.get('code', '') or '').strip()
    new_password = str(data.get('new_password', '') or '')

    if not all([national_id, role, code, new_password]):
        return jsonify({'error': 'national_id, role, code, and new_password are required'}), 400
    if len(new_password) < 8:
        return jsonify({'error': 'Password must be at least 8 characters'}), 400

    if role == 'admin':
        user = Admin.query.filter_by(national_id=national_id).first()
    elif role == 'specialist':
        user = Specialist.query.filter_by(national_id=national_id).first()
    elif role == 'RegisteredUser':
        user = RegisteredUser.query.filter_by(national_id=national_id).first()
    else:
        return jsonify({'error': 'Invalid role'}), 400

    if not user:
        return jsonify({'error': 'Invalid ID or code'}), 400

    if not user.verification_code or user.verification_code != code:
        return jsonify({'error': 'Invalid or expired reset code'}), 400

    if user.verification_expires and datetime.utcnow() > user.verification_expires:
        return jsonify({'error': 'Reset code has expired. Request a new one.'}), 400

    user.password             = generate_password_hash(new_password)
    user.verification_code    = None
    user.verification_expires = None
    db.session.commit()

    print(f"✅ Password reset for national_id={national_id}")
    return jsonify({'message': 'Password updated successfully. You can now log in.'}), 200

# ══════════════════════════════════════════════════════════
#   ADMIN 
# ══════════════════════════════════════════════════════════


##======    Admin: Get all users   ======##
@app.route('/admin/users', methods=['GET'])
def get_all_users():
    admins      = Admin.query.all()
    specialists = Specialist.query.all()
    registeredUser    = RegisteredUser.query.all()

    users = []
    for a in admins:
        users.append({'id': a.national_id, 'name': a.name, 'email': a.email,
                      'role': 'Admin', 'status': 'active', 'nationalId': a.national_id,
                      'gender': (a.gender or '')})
    for s in specialists:
        users.append({'id': s.national_id, 'name': s.name, 'email': s.email,
                      'role': 'Specialist', 'status': 'active', 'nationalId': s.national_id,
                      'gender': (s.gender or '')})
    for p in registeredUser:
        users.append({'id': p.national_id, 'name': p.name or 'Pending', 'email': p.email,
                      'role': 'Registered User', 'status': 'active', 'nationalId': p.national_id, 'gender': p.gender or ''})

    print(f"📋 Returning {len(users)} users")
    return jsonify(users), 200


##======    Admin: Add user   ======##
@app.route('/admin/users', methods=['POST'])
def add_user():
    data        = request.get_json()
    national_id = data.get('national_id', '').strip()
    name        = data.get('name', '').strip()
    email       = data.get('email', '').strip()
    role        = data.get('role', '')
    phone       = (data.get('phone') or data.get('phone_num') or '').strip()
    gender_raw  = (data.get('gender') or '').strip()
    temp_password = 'Temp@1234'
    performed_by_name = data.get('performed_by_name', 'Admin')
    performed_by_id   = data.get('performed_by_id', 'unknown')
    tables =[Patient,Admin, Specialist, RegisteredUser]


    if gender_raw not in ('Male', 'Female', ''):
        return jsonify({'error': 'Gender must be Male or Female'}), 400
    gender_store = gender_raw if gender_raw in ('Male', 'Female') else None

    if not all([national_id, name, email, role]):
        return jsonify({'error': 'All fields are required'}), 400

    try:
        if role == 'Admin':
            if Admin.query.filter_by(national_id=national_id).first():
                return jsonify({'error': 'National ID already exists'}), 409
            if Admin.query.filter_by(email=email).first():
                return jsonify({'error': 'Email already exists'}), 409
            db.session.add(Admin(
                national_id=national_id,
                name=name,
                email=email,
                password=generate_password_hash(temp_password),
                gender=gender_store,
            ))
        elif role == 'Specialist':
            if Specialist.query.filter_by(national_id=national_id).first():
                return jsonify({'error': 'National ID already exists'}), 409
            if Specialist.query.filter_by(email=email).first():
                return jsonify({'error': 'Email already exists'}), 409
            db.session.add(Specialist(
                national_id=national_id,
                name=name,
                email=email,
                password=generate_password_hash(temp_password),
                gender=gender_store,
            ))
        elif role in ('Registered User', 'Patient'):
            if not phone:
                return jsonify({'error': 'Phone number is required for registered users'}), 400
            if gender_store not in ('Male', 'Female'):
                return jsonify({'error': 'Gender is required for registered users'}), 400
            
            for table in tables:
                if table.query.filter_by(national_id=national_id).first():
                    return jsonify({'error': 'National ID already registered'}), 409
        
            if RegisteredUser.query.filter_by(email=email).first():
                return jsonify({'error': 'Email already exists'}), 409
            if RegisteredUser.query.filter_by(phone_num=phone).first():
                return jsonify({'error': 'Phone number already exists'}), 409
            db.session.add(RegisteredUser(
                national_id=national_id,
                name=name,
                email=email,
                phone_num=phone,
                gender=gender_store,
                password=generate_password_hash(temp_password),
                is_verified=True,
                verification_code=None,
                verification_expires=None,
            ))
            role = 'Registered User'
        else:
            return jsonify({'error': f'Unknown role: {role}'}), 400

        db.session.commit()
        print(f"✅ Admin added {role}: {name}")
        write_system_log(performed_by_name, performed_by_id, f'Added new {role}: {name} — ID: {national_id}', role='Admin')
        return jsonify({
            'message': f'{role} added successfully',
            'temp_password': temp_password,
        }), 201

    except Exception as e:
        db.session.rollback()
        print(f"❌ Error: {e}")
        return jsonify({'error': str(e)}), 500


##======    Admin: Delete user   ======##
@app.route('/admin/users/<national_id>', methods=['DELETE'])
def delete_user(national_id):
    from models import EEGSession, Alert
    role = request.args.get('role', '')
    data              = request.get_json(silent=True) or {}
    performed_by_name = data.get('performed_by_name', 'Admin')
    performed_by_id   = data.get('performed_by_id', 'unknown')

    if role == 'Admin':
        user = Admin.query.filter_by(national_id=national_id).first()
    elif role == 'Specialist':
        user = Specialist.query.filter_by(national_id=national_id).first()
    else:
        user = RegisteredUser.query.filter_by(national_id=national_id).first()

    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    if role == 'Admin' and national_id == performed_by_id:
        return jsonify({'error': 'You cannot delete your own account.'}), 403

    # If deleting a specialist, clean up their related records first
    # (because eeg_session.specialist_national_id is NOT NULL)
    if role == 'Specialist':
        sessions = EEGSession.query.filter_by(specialist_national_id=national_id).all()
        for session in sessions:
            # Delete alerts linked to this session first
            Alert.query.filter_by(alert_session_id=session.session_id).delete()
        # Now delete the sessions
        EEGSession.query.filter_by(specialist_national_id=national_id).delete()
        # Unlink their patients (set specialist to None — patient stays, just unassigned)
        Patient.query.filter_by(specialist_national_id=national_id).update(
            {'specialist_national_id': None}
        )

    db.session.delete(user)
    db.session.commit()
    print(f"🗑️ Deleted {role}: {national_id}")
    write_system_log(performed_by_name, performed_by_id, f'Deleted user: {user.name} — ID: {national_id}', role='Admin')
    return jsonify({'message': 'User deleted successfully'}), 200

##======    Admin: Edit user   ======##
@app.route('/admin/users/<national_id>', methods=['PUT'])
def edit_user(national_id):
    data  = request.get_json()
    name  = data.get('name', '').strip()
    email = data.get('email', '').strip()
    role  = data.get('role', '')
    performed_by_name = data.get('performed_by_name', 'Admin')
    performed_by_id   = data.get('performed_by_id', 'unknown')

    if not all([name, email]):
        return jsonify({'error': 'Name and email are required'}), 400
    if role == 'Admin':
        user = Admin.query.filter_by(national_id=national_id).first()
    elif role == 'Specialist':
        user = Specialist.query.filter_by(national_id=national_id).first()
    else:
        user = RegisteredUser.query.filter_by(national_id=national_id).first()
    if not user:
        return jsonify({'error': 'User not found'}), 404
    user.name  = name
    user.email = email
    db.session.commit()
    write_system_log(performed_by_name, performed_by_id, f'Edited user info: {user.name} — ID: {national_id}', role='Admin')
    return jsonify({'message': 'User updated successfully'}), 200

def _current_production_model():
    """Prefer latest Active model; otherwise latest row by id."""
    m = Model.query.filter_by(model_status='Active').order_by(Model.model_id.desc()).first()
    if not m:
        m = Model.query.order_by(Model.model_id.desc()).first()
    return m


##======    Admin: Current model metadata (for Models page)   ======##
@app.route('/admin/current-model', methods=['GET'])
def get_current_model_info():
    m = _current_production_model()
    if not m:
        return jsonify({
            'model_id': None,
            'model_name': '',
            'model_version': '',
            'training_date': '',
            'model_accuracy': None,
            'model_status': '',
        }), 200
    td = m.training_date
    date_str = td.date().isoformat() if td else ''
    acc = float(m.model_accuracy) if m.model_accuracy is not None else None
    return jsonify({
        'model_id': m.model_id,
        'model_name': m.model_name or '',
        'model_version': m.model_version or '',
        'training_date': date_str,
        'model_accuracy': acc,
        'model_status': str(m.model_status or ''),
    }), 200


@app.route('/admin/models', methods=['GET'])
def admin_list_models():
    """All decoder model rows (for admin Models performance chart)."""
    rows = Model.query.order_by(Model.training_date.asc(), Model.model_id.asc()).all()
    out = []
    for m in rows:
        td = m.training_date
        date_str = td.date().isoformat() if td else ''
        out.append({
            'model_id': m.model_id,
            'model_name': m.model_name or '',
            'model_version': m.model_version or '',
            'model_accuracy': float(m.model_accuracy) if m.model_accuracy is not None else 0.0,
            'model_status': str(m.model_status or ''),
            'training_date': date_str,
        })
    return jsonify({'models': out}), 200


@app.route('/admin/current-model', methods=['PUT'])
def put_current_model_info():
    data = request.get_json() or {}
    name = data.get('model_name', '').strip()
    version = data.get('model_version', '').strip()
    date_str = data.get('training_date', '').strip()
    performed_by_name = data.get('performed_by_name', 'Admin')
    performed_by_id = str(data.get('performed_by_id', '') or '').strip() or 'unknown'

    if not name or not version or not date_str:
        return jsonify({'error': 'Model name, version, and training date are required'}), 400
    try:
        parsed = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Training date must be YYYY-MM-DD'}), 400
    training_dt = datetime.combine(parsed, datetime.min.time())

    admin_row = Admin.query.filter_by(national_id=performed_by_id).first()
    if not admin_row:
        admin_row = Admin.query.first()
    if not admin_row:
        return jsonify({'error': 'No admin account available to attach model'}), 500
    admin_nat = admin_row.national_id

    m = _current_production_model()
    try:
        if m:
            m.model_name = name
            m.model_version = version
            m.training_date = training_dt
        else:
            db.session.add(Model(
                model_name=name,
                model_version=version,
                model_accuracy=Decimal('0.9500'),
                model_status='Active',
                training_date=training_dt,
                admin_national_id=admin_nat,
            ))
        db.session.commit()
        write_system_log(
            performed_by_name,
            performed_by_id,
            f'Updated production model: {name} ({version}), trained {date_str}',
            role='Admin',
        )
        return jsonify({'message': 'Model information saved'}), 200
    except Exception as e:
        db.session.rollback()
        print(f'❌ Model update error: {e}')
        return jsonify({'error': str(e)}), 500


##======    Admin: System Logs   ======##
@app.route('/admin/system-logs', methods=['GET'])
def get_system_logs():
    logs = SystemLog.query.order_by(SystemLog.log_timestamp.desc()).limit(100).all()
    result = []
    for log in logs:
        result.append({
            'id':          log.log_id,
            'userName':    log.user_name,
            'role':        log.role or '—',
            'nationalId':  log.user_national_id,
            'event':       log.event,
            'timestamp':   log.log_timestamp.strftime('%Y-%m-%d %H:%M:%S'),
        })
    return jsonify(result), 200


# ══════════════════════════════════════════════════════════
#   ADMIN: Model artifact status + sessions
# ══════════════════════════════════════════════════════════

@app.route('/admin/model-artifact-status', methods=['GET'])
def admin_model_artifact_status():
    repo_root = Path(BASE_DIR).parent
    saved_dir = repo_root / 'Saved_Model'
    pkl_path = saved_dir / 'natiqi_5word_svm_pipeline.pkl'
    config_path = saved_dir / 'config.json'
    metadata_path = saved_dir / 'metadata.json'

    payload = {
        'pkl_exists': pkl_path.exists(),
        'pkl_path': str(pkl_path),
        'pkl_size_bytes': int(pkl_path.stat().st_size) if pkl_path.exists() else 0,
        'pkl_modified_at': datetime.fromtimestamp(pkl_path.stat().st_mtime).isoformat() if pkl_path.exists() else '',
        'config': {},
        'metadata': {},
    }
    try:
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                payload['config'] = json.load(f)
        if metadata_path.exists():
            with open(metadata_path, 'r', encoding='utf-8') as f:
                payload['metadata'] = json.load(f)
    except Exception as e:
        payload['read_error'] = str(e)

    return jsonify(payload), 200


def _compute_top_predicted_word_from_events(events):
    """Most frequent detected_word in events; avg confidence only for rows that match that word."""
    counts = {}
    sum_conf = {}
    n_conf = {}
    for e in events:
        w = (getattr(e, 'detected_word', None) or '').strip()
        if not w:
            continue
        counts[w] = counts.get(w, 0) + 1
        if getattr(e, 'confidence', None) is not None:
            cf = float(e.confidence)
            sum_conf[w] = sum_conf.get(w, 0.0) + cf
            n_conf[w] = n_conf.get(w, 0) + 1
    if not counts:
        return None, None
    top_word = max(counts.keys(), key=lambda w: (counts[w], w))
    if n_conf.get(top_word, 0) <= 0:
        return top_word, None
    return top_word, sum_conf[top_word] / n_conf[top_word]


def _serialize_eeg_session_list(sessions):
    """Bulk-load events and attach top_predicted_word + per-word avg confidence."""
    out = []
    if not sessions:
        return out
    ids = [s.session_id for s in sessions]
    ev_by = {}
    if ids:
        for ev in EEGSessionEvent.query.filter(EEGSessionEvent.session_id.in_(ids)).all():
            ev_by.setdefault(ev.session_id, []).append(ev)
    for s in sessions:
        evs = ev_by.get(s.session_id, [])
        top_w, top_avg = _compute_top_predicted_word_from_events(evs)
        if not evs:
            top_w = s.detected_word or ''
            top_avg = float(s.confidence_level) if s.confidence_level is not None else None
        else:
            if not top_w:
                top_w = s.detected_word or ''
        out.append({
            'session_id': s.session_id,
            'patient_national_id': str(s.patient_national_id),
            'specialist_national_id': (str(s.specialist_national_id) if s.specialist_national_id else ''),
            'model_id': s.model_id,
            'start_time': s.start_time.isoformat() if s.start_time else '',
            'end_time': s.end_time.isoformat() if s.end_time else '',
            'detected_word': s.detected_word,
            'confidence_level': float(s.confidence_level) if s.confidence_level is not None else None,
            'top_predicted_word': top_w,
            'top_predicted_word_avg_confidence': top_avg,
            'device': s.device or '',
            'channels': int(s.channels) if s.channels is not None else 14,
            'session_status': s.session_status,
        })
    return out


@app.route('/admin/sessions', methods=['GET'])
def admin_list_sessions():
    limit = request.args.get('limit', default=100, type=int)
    q = EEGSession.query.order_by(EEGSession.session_id.desc()).limit(max(1, min(limit, 500))).all()
    return jsonify(_serialize_eeg_session_list(q)), 200


# ══════════════════════════════════════════════════════════
#   ML (EEG inference)
# ══════════════════════════════════════════════════════════

@app.route('/ml/predict-window', methods=['POST'])
def ml_predict_window():
    if predict_window is None:
        return jsonify({'error': 'ML module not available on server'}), 500

    data = request.get_json(silent=True) or {}
    window = data.get('window', None)
    if window is None:
        return jsonify({'error': 'Missing required field: window'}), 400

    try:
        # Expect 14x128 numeric matrix
        pred = predict_window(window)
        return jsonify(pred.to_json()), 200
    except ModelNotReadyError as e:
        return jsonify({'error': str(e)}), 503
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback
        print('❌ /ml/predict-window error:')
        traceback.print_exc()
        return jsonify({'error': 'Inference failed'}), 500


@app.route('/ml/live-demo/predict', methods=['POST'])
def ml_live_demo_predict():
    """
    Demo-only: predicts one of the 4 critical words (جوع/عطش/حمام/دواء) using the
    highest-accuracy RF artifacts under LiveDataModels, sampling from cleaned test rows.
    """
    if predict_live_demo is None:
        return jsonify({'error': 'Live demo ML module not available on server'}), 500

    data = request.get_json(silent=True) or {}
    subject = str(data.get('subject') or 'aya').strip() or 'aya'
    seed = data.get('seed', None)

    try:
        pred = predict_live_demo(subject=subject, seed=seed)
        return jsonify(pred.to_json()), 200
    except LiveDemoModelNotReadyError as e:
        return jsonify({'error': str(e)}), 503
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception:
        import traceback
        print('❌ /ml/live-demo/predict error:')
        traceback.print_exc()
        return jsonify({'error': 'Inference failed'}), 500


@app.route('/eeg/sessions', methods=['POST'])
def create_eeg_session_from_window():
    """
    Create an EEGSession row by running ML inference on a provided 14x128 window.
    This is a starter endpoint for integration/testing; real devices will likely
    send multiple windows per session later.
    """
    if predict_window is None:
        return jsonify({'error': 'ML module not available on server'}), 500

    data = request.get_json(silent=True) or {}
    patient_national_id = str(data.get('patient_national_id', '')).strip()
    specialist_national_id = str(data.get('specialist_national_id', '')).strip() or None
    window = data.get('window', None)
    device = (data.get('device') or 'EPOC X').strip()

    if not patient_national_id or window is None:
        return jsonify({'error': 'patient_national_id and window are required'}), 400

    patient = Patient.query.filter_by(national_id=patient_national_id).first()
    registered_user = RegisteredUser.query.filter_by(national_id=patient_national_id).first()

    if not patient and not registered_user:
        return jsonify({'error': 'Patient not found'}), 404

    patient_id_to_use = patient_national_id  

    specialist_row = None
    if specialist_national_id:
        specialist_row = Specialist.query.filter_by(national_id=specialist_national_id).first()

    try:
        pred = predict_window(window)
    except ModelNotReadyError as e:
        return jsonify({'error': str(e)}), 503
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        print(f'❌ /eeg/sessions inference error: {e}')
        return jsonify({'error': 'Inference failed'}), 500

    now = datetime.utcnow()
    # Attach current production model row if available
    m = _current_production_model()
    try:
        row = EEGSession(
            patient_national_id=patient_id_to_use,
            specialist_national_id=(specialist_row.national_id if specialist_row else None),
            model_id=(m.model_id if m else None),
            start_time=now,
            end_time=now,
            detected_word=pred.predicted_word_ar,
            confidence_level=Decimal(str(pred.confidence)),
            device=device,
            channels=14,
            session_status='Ended',
        )
        db.session.add(row)
        db.session.commit()
        return jsonify({
            'message': 'EEG session created',
            'session_id': row.session_id,
            'prediction': pred.to_json(),
        }), 201
    except Exception as e:
        db.session.rollback()
        print(f'❌ /eeg/sessions db error: {e}')
        return jsonify({'error': 'Failed to save session'}), 500


@app.route('/eeg/live-demo/sessions', methods=['POST'])
def create_live_demo_session():
    """
    Demo-only: persist a live-demo EEGSession row (predicted word/confidence + time range).
    Used by the recipient dashboard when pressing Stop.
    """
    data = request.get_json(silent=True) or {}
    patient_national_id = str(data.get('patient_national_id', '')).strip()
    device = (data.get('device') or 'EPOC X').strip()
    detected_word = str(data.get('detected_word') or '').strip()
    confidence = data.get('confidence', None)
    events = data.get('events', None)  # [{event_time, detected_word, confidence}]
    start_time = str(data.get('start_time') or '').strip()
    end_time = str(data.get('end_time') or '').strip()

    if not patient_national_id or not detected_word or confidence is None:
        return jsonify({'error': 'patient_national_id, detected_word and confidence are required'}), 400

    patient = Patient.query.filter_by(national_id=patient_national_id).first()
    registered_user = RegisteredUser.query.filter_by(national_id=patient_national_id).first()
    if not patient and not registered_user:
        return jsonify({'error': 'Patient not found'}), 404

    try:
        conf_f = float(confidence)
    except Exception:
        return jsonify({'error': 'confidence must be a number'}), 400

    def _parse_dt(s: str) -> datetime:
        if not s:
            return datetime.utcnow()
        try:
            # Accept ISO strings (with or without timezone 'Z')
            return datetime.fromisoformat(s.replace('Z', '+00:00')).replace(tzinfo=None)
        except Exception:
            return datetime.utcnow()

    st = _parse_dt(start_time)
    et = _parse_dt(end_time) if end_time else datetime.utcnow()
    m = _current_production_model()

    # If we have events, compute summary fields (duration, avg confidence, most repeated word)
    most_word = detected_word
    avg_conf = conf_f
    parsed_events: list[dict] = []
    if isinstance(events, list) and len(events) > 0:
        word_counts: dict[str, int] = {}
        conf_sum = 0.0
        conf_n = 0
        for ev in events:
            if not isinstance(ev, dict):
                continue
            w = str(ev.get('detected_word') or '').strip()
            if not w:
                continue
            c = ev.get('confidence', None)
            try:
                c_f = float(c) if c is not None else None
            except Exception:
                c_f = None
            t = _parse_dt(str(ev.get('event_time') or ''))
            parsed_events.append({'event_time': t, 'detected_word': w, 'confidence': c_f})
            word_counts[w] = word_counts.get(w, 0) + 1
            if c_f is not None:
                conf_sum += c_f
                conf_n += 1
        if word_counts:
            most_word = max(word_counts.items(), key=lambda kv: kv[1])[0]
        if conf_n > 0:
            avg_conf = conf_sum / conf_n

    try:
        row = EEGSession(
            patient_national_id=patient_national_id,
            specialist_national_id=None,
            model_id=(m.model_id if m else None),
            start_time=st,
            end_time=et,
            detected_word=most_word,
            confidence_level=Decimal(str(avg_conf)),
            device=device,
            channels=14,
            session_status='Ended',
        )
        db.session.add(row)
        db.session.commit()

        # Persist events if provided
        try:
            for ev in parsed_events:
                db.session.add(EEGSessionEvent(
                    session_id=row.session_id,
                    event_time=ev['event_time'],
                    detected_word=ev['detected_word'],
                    confidence=(Decimal(str(ev['confidence'])) if ev.get('confidence') is not None else None),
                ))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f'⚠️ live-demo events save failed (session_id={row.session_id}): {e}')

        return jsonify({'message': 'Live demo session saved', 'session_id': row.session_id}), 201
    except Exception as e:
        db.session.rollback()
        print(f'❌ /eeg/live-demo/sessions db error: {e}')
        return jsonify({'error': 'Failed to save live demo session'}), 500


@app.route('/eeg/live-demo/sessions/<int:session_id>/report', methods=['GET'])
def live_demo_session_report(session_id: int):
    s = EEGSession.query.filter_by(session_id=session_id).first()
    if not s:
        return jsonify({'error': 'Session not found'}), 404

    evs = EEGSessionEvent.query.filter_by(session_id=session_id).order_by(EEGSessionEvent.event_time.asc()).all()
    events_out = []
    word_counts: dict[str, int] = {}
    conf_sum = 0.0
    conf_n = 0

    st = s.start_time or datetime.utcnow()
    for e in evs:
        word_counts[e.detected_word] = word_counts.get(e.detected_word, 0) + 1
        c = float(e.confidence) if e.confidence is not None else None
        if c is not None:
            conf_sum += c
            conf_n += 1
        elapsed = int((e.event_time - st).total_seconds())
        h = elapsed // 3600
        m = (elapsed % 3600) // 60
        sec = elapsed % 60
        events_out.append({
            'event_time': e.event_time.isoformat(),
            'day': e.event_time.date().isoformat(),
            'elapsed': f'{h:02d}:{m:02d}:{sec:02d}',
            'detected_word': e.detected_word,
            'confidence': c,
        })

    duration_sec = int((s.end_time - st).total_seconds()) if s.end_time else 0
    avg_conf = (conf_sum / conf_n) if conf_n > 0 else (float(s.confidence_level) if s.confidence_level is not None else None)
    most_word = max(word_counts.items(), key=lambda kv: kv[1])[0] if word_counts else s.detected_word

    return jsonify({
        'session_id': s.session_id,
        'start_time': s.start_time.isoformat() if s.start_time else '',
        'end_time': s.end_time.isoformat() if s.end_time else '',
        'duration_seconds': duration_sec,
        'avg_confidence': avg_conf,
        'most_repeated_word': most_word,
        'word_counts': word_counts,
        'events': events_out,
    }), 200


@app.route('/eeg/live-demo/sessions/<int:session_id>/report.csv', methods=['GET'])
def live_demo_session_report_csv(session_id: int):
    s = EEGSession.query.filter_by(session_id=session_id).first()
    if not s:
        return jsonify({'error': 'Session not found'}), 404

    evs = EEGSessionEvent.query.filter_by(session_id=session_id).order_by(EEGSessionEvent.event_time.asc()).all()
    def _fmt_dt(dt):
        if not dt:
            return ''
        try:
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            try:
                return dt.isoformat()
            except Exception:
                return ''

    output = io.StringIO()
    writer = csv.writer(output)

    # Header summary
    writer.writerow(['session_id', session_id])
    writer.writerow(['start_time', _fmt_dt(s.start_time)])
    writer.writerow(['end_time', _fmt_dt(s.end_time)])
    try:
        duration_sec = int((s.end_time - s.start_time).total_seconds()) if s.start_time and s.end_time else 0
    except Exception:
        duration_sec = 0
    writer.writerow(['duration_seconds', duration_sec])
    writer.writerow(['avg_confidence', float(s.confidence_level) if s.confidence_level is not None else ''])
    writer.writerow([])

    # Events table
    writer.writerow(['elapsed', 'day', 'event_time', 'detected_word', 'confidence_percent'])
    st = s.start_time or datetime.utcnow()
    for e in evs:
        elapsed = int((e.event_time - st).total_seconds())
        h = elapsed // 3600
        m = (elapsed % 3600) // 60
        sec = elapsed % 60
        c = float(e.confidence) if e.confidence is not None else None
        writer.writerow([
            f'{h:02d}:{m:02d}:{sec:02d}',
            e.event_time.date().isoformat() if getattr(e, 'event_time', None) else '',
            _fmt_dt(getattr(e, 'event_time', None)),
            e.detected_word,
            (round(c * 100, 2) if c is not None else ''),
        ])

    # Add UTF-8 BOM so Excel reads Arabic correctly.
    csv_data = '\ufeff' + output.getvalue()
    return Response(
        csv_data,
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename=\"rs-{session_id}.csv\"'},
    )


@app.route('/eeg/live-demo/sessions/<int:session_id>/report.xlsx', methods=['GET'])
def live_demo_session_report_xlsx(session_id: int):
    s = EEGSession.query.filter_by(session_id=session_id).first()
    if not s:
        return jsonify({'error': 'Session not found'}), 404

    evs = EEGSessionEvent.query.filter_by(session_id=session_id).order_by(EEGSessionEvent.event_time.asc()).all()

    def _fmt_dt(dt):
        if not dt:
            return ''
        try:
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            try:
                return dt.isoformat()
            except Exception:
                return ''

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception as e:
        return jsonify({'error': f'Excel export not available: {e}'}), 500

    wb = Workbook()
    # Rename default sheet to Session Info
    ws_info = wb.active
    ws_info.title = 'Session Info'
    ws_pred = wb.create_sheet('Predictions')
    ws_sum = wb.create_sheet('Summary')

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F2937')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _write_kv_table(ws, rows):
        ws.append(['Field', 'Value'])
        for r in ws[1]:
            r.font = header_font
            r.fill = header_fill
            r.alignment = header_align
        for k, v in rows:
            ws.append([k, v])
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 44
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
            for c in row:
                c.alignment = cell_align

    try:
        duration_sec = int((s.end_time - s.start_time).total_seconds()) if s.start_time and s.end_time else 0
    except Exception:
        duration_sec = 0
    duration_label = f'{duration_sec // 3600:02d}:{(duration_sec % 3600) // 60:02d}:{duration_sec % 60:02d}'

    _write_kv_table(ws_info, [
        ('Session ID', s.session_id),
        ('Start', _fmt_dt(s.start_time)),
        ('End', _fmt_dt(s.end_time)),
        ('Duration', duration_label),
        ('Avg confidence', (float(s.confidence_level) if s.confidence_level is not None else '')),
    ])

    # Predictions sheet
    pred_header = ['#', 'Elapsed', 'Day', 'Time', 'Predicted word', 'Confidence %']
    ws_pred.append(pred_header)
    for r in ws_pred[1]:
        r.font = header_font
        r.fill = header_fill
        r.alignment = header_align

    st = s.start_time or datetime.utcnow()
    for idx, e in enumerate(evs, start=1):
        try:
            elapsed = int((e.event_time - st).total_seconds())
        except Exception:
            elapsed = 0
        h = elapsed // 3600
        m = (elapsed % 3600) // 60
        sec = elapsed % 60
        c = float(e.confidence) if e.confidence is not None else None
        ws_pred.append([
            idx,
            f'{h:02d}:{m:02d}:{sec:02d}',
            e.event_time.date().isoformat() if getattr(e, 'event_time', None) else '',
            _fmt_dt(getattr(e, 'event_time', None)),
            e.detected_word,
            (round(c * 100, 2) if c is not None else ''),
        ])

    # Column widths
    widths = [6, 12, 12, 20, 22, 14]
    for i, w in enumerate(widths, start=1):
        ws_pred.column_dimensions[get_column_letter(i)].width = w
    for row in ws_pred.iter_rows(min_row=2, max_row=ws_pred.max_row, min_col=1, max_col=len(pred_header)):
        for c in row:
            c.alignment = cell_align

    # Summary sheet
    word_counts = {}
    conf_sum = 0.0
    conf_n = 0
    for e in evs:
        word_counts[e.detected_word] = word_counts.get(e.detected_word, 0) + 1
        c = float(e.confidence) if e.confidence is not None else None
        if c is not None:
            conf_sum += c
            conf_n += 1

    most_word = max(word_counts.items(), key=lambda kv: kv[1])[0] if word_counts else (s.detected_word or '')
    most_count = int(word_counts.get(most_word, 0)) if most_word else 0
    avg_conf = (conf_sum / conf_n) if conf_n > 0 else (float(s.confidence_level) if s.confidence_level is not None else None)

    ws_sum.append(['Metric', 'Value'])
    for r in ws_sum[1]:
        r.font = header_font
        r.fill = header_fill
        r.alignment = header_align
    ws_sum.append(['Total predictions', len(evs)])
    ws_sum.append(['Most repeated word', most_word])
    ws_sum.append(['Most repeated count', most_count])
    ws_sum.append(['Avg confidence', (round(avg_conf * 100, 2) if avg_conf is not None else '')])
    ws_sum.append([])
    ws_sum.append(['Word', 'Count'])
    for r in ws_sum[ws_sum.max_row]:
        r.font = header_font
        r.fill = header_fill
        r.alignment = header_align
    for w, n in sorted(word_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws_sum.append([w, int(n)])

    ws_sum.column_dimensions['A'].width = 26
    ws_sum.column_dimensions['B'].width = 20
    for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row, min_col=1, max_col=2):
        for c in row:
            c.alignment = cell_align

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        bio.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=\"rs-{session_id}.xlsx\"'},
    )


# ══════════════════════════════════════════════════════════
#   SPECIALIST: Sessions
# ══════════════════════════════════════════════════════════

@app.route('/specialist/sessions', methods=['GET'])
def specialist_list_sessions():
    specialist_id = request.args.get('specialist_id', default='', type=str).strip()
    patient_id = request.args.get('patient_national_id', default='', type=str).strip()
    limit = request.args.get('limit', default=100, type=int)

    q = EEGSession.query
    if specialist_id:
        q = q.filter(EEGSession.specialist_national_id == specialist_id)
    if patient_id:
        q = q.filter(EEGSession.patient_national_id == patient_id)

    sessions = q.order_by(EEGSession.session_id.desc()).limit(max(1, min(limit, 500))).all()
    return jsonify(_serialize_eeg_session_list(sessions)), 200


# ══════════════════════════════════════════════════════════
#   SPECIALIST 
# ══════════════════════════════════════════════════════════
##======    Specialist: Get patients   ======##
@app.route('/specialist/patients', methods=['GET'])
def get_patients():
    # Optional: filter by specialist's national_id
    specialist_id = request.args.get('specialist_id')

    if specialist_id:
        patients = Patient.query.filter_by(specialist_national_id=specialist_id).all()
    else:
        patients = Patient.query.all()

    result = []
    for p in patients:
        result.append({
            'id':          p.national_id,
            'nationalId':  p.national_id,
            'roomNumber':  (p.room_number or '').strip(),
            'name':        p.name,
            'dob':         p.date_of_birth.strftime('%Y-%m-%d') if p.date_of_birth else '',
            'gender':      p.gender or '',
            'device':      p.device or 'Emotiv EPOC X',
            'status':      p.status or 'Active',
        })

    print(f"📋 Returning {len(result)} patients")
    return jsonify(result), 200


##======    Specialist: Add patient   ======##
@app.route('/specialist/patients', methods=['POST'])
def add_patient():
    data          = request.get_json()
    national_id   = data.get('national_id', '').strip()
    room_number   = data.get('room_number', '').strip()
    name          = data.get('name', '').strip()
    dob           = data.get('dob', '').strip()       # format: YYYY-MM-DD
    gender        = data.get('gender', '').strip()
    specialist_id = data.get('specialist_id', '').strip()
    performed_by_name = data.get('performed_by_name', 'Specialist')
    performed_by_id   = data.get('performed_by_id', 'unknown')
    tables =[Patient,Admin, Specialist, RegisteredUser]

    if not all([national_id, room_number, name, dob]):
        return jsonify({'error': 'Room number, name, National ID, and DOB are required'}), 400
    if len(room_number) > 10:
        return jsonify({'error': 'Room number must be at most 10 characters'}), 400

    for table in tables:
        if table.query.filter_by(national_id=national_id).first():
            return jsonify({'error': 'National ID already registered'}), 409
        
    if Patient.query.filter_by(room_number=room_number).first():
        return jsonify({'error': 'Room number already in use'}), 409

    try:
        from datetime import date
        dob_parsed = date.fromisoformat(dob)   # validates YYYY-MM-DD format

        new_patient = Patient(
            national_id            = national_id,
            name                   = name,
            password               = generate_password_hash('user@123'),
            date_of_birth          = dob_parsed,
            gender                 = gender,
            room_number            = room_number,
            device                 = 'EPOC X',
            status                 = 'Active',
            specialist_national_id = specialist_id or None,
        )
        db.session.add(new_patient)
        db.session.commit()

        print(f"✅ Added patient: {name}")
        write_system_log(performed_by_name, performed_by_id, f'Added new patient: {name} — ID: {national_id}', role='Specialist')
        return jsonify({'message': 'Patient added successfully'}), 201

    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error: {e}")
        return jsonify({'error': str(e)}), 500


##======    Specialist: Delete patient   ======##
@app.route('/specialist/patients/<national_id>', methods=['DELETE'])
def delete_patient(national_id):
    patient = Patient.query.filter_by(national_id=national_id).first()
    data              = request.get_json(silent=True) or {}
    performed_by_name = data.get('performed_by_name', 'Specialist')
    performed_by_id   = data.get('performed_by_id', 'unknown')

    if not patient:
        return jsonify({'error': 'Patient not found'}), 404

    try:
        # Delete alerts linked to this patient's sessions first
        sessions = EEGSession.query.filter_by(patient_national_id=national_id).all()
        for session in sessions:
            Alert.query.filter_by(alert_session_id=session.session_id).delete()

        # Delete the sessions themselves
        EEGSession.query.filter_by(patient_national_id=national_id).delete()

        # Delete patient settings
        PatientSettings.query.filter_by(user_national_id=national_id).delete()

        # Now safe to delete the patient
        db.session.delete(patient)
        db.session.commit()

        print(f"🗑️ Deleted patient: {national_id}")
        write_system_log(performed_by_name, performed_by_id, f'Deleted patient: {patient.name} — ID: {national_id}', role='Specialist')
        return jsonify({'message': 'Patient deleted successfully'}), 200

    except Exception as e:
        db.session.rollback()
        print(f"❌ Delete patient error: {e}")
        return jsonify({'error': str(e)}), 500

##======    Specialist: Edit patient   ======##
@app.route('/specialist/patients/<national_id>', methods=['PUT'])
def edit_patient(national_id):
    data   = request.get_json()
    room_number = data.get('room_number', '').strip()
    name   = data.get('name', '').strip()
    dob    = data.get('dob', '').strip()
    gender = data.get('gender', '').strip()
    performed_by_name = data.get('performed_by_name', 'Specialist')
    performed_by_id   = data.get('performed_by_id', 'unknown')

    if not all([room_number, name, dob]):
        return jsonify({'error': 'Room number, name, and DOB are required'}), 400
    if len(room_number) > 10:
        return jsonify({'error': 'Room number must be at most 10 characters'}), 400
    patient = Patient.query.filter_by(national_id=national_id).first()
    if not patient:
        return jsonify({'error': 'Patient not found'}), 404
    taken = Patient.query.filter(
        Patient.room_number == room_number,
        Patient.national_id != national_id,
    ).first()
    if taken:
        return jsonify({'error': 'Room number already in use'}), 409
    try:
        from datetime import date
        patient.room_number   = room_number
        patient.name          = name
        patient.date_of_birth = date.fromisoformat(dob)
        patient.gender        = gender
        db.session.commit()
        write_system_log(performed_by_name, performed_by_id, f'Edited patient info: {patient.name} — ID: {national_id}', role='Specialist')
        return jsonify({'message': 'Patient updated successfully'}), 200
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

##======    Specialist: Toggle patient status   ======##
@app.route('/specialist/patients/<national_id>/status', methods=['PUT'])
def toggle_patient_status(national_id):
    data              = request.get_json(silent=True) or {}
    performed_by_name = data.get('performed_by_name', 'Specialist')
    performed_by_id   = data.get('performed_by_id', 'unknown')

    patient = Patient.query.filter_by(national_id=national_id).first()

    if not patient:
        return jsonify({'error': 'Patient not found'}), 404

    # Toggle between stable/active and suspended
    if patient.status ==  'Active':
        patient.status = 'Inactive'
    else:
        patient.status = 'Active'

    db.session.commit()
    print(f"⏸️ Toggled patient {national_id} status to: {patient.status}")
    write_system_log(performed_by_name, performed_by_id, f'Toggled patient status: {patient.name} — ID: {national_id}', role='Specialist')
    return jsonify({'message': 'Status updated', 'status': patient.status}), 200



# ══════════════════════════════════════════════════════════
#   REGISTRED USER 
# ══════════════════════════════════════════════════════════
##======    Update user profile   ======##
@app.route('/profile/update', methods=['PUT'])
def update_profile():
    data        = request.get_json()
    national_id = data.get('national_id', '').strip()
    role        = data.get('role', '')
    name        = data.get('name', '').strip()
    email       = data.get('email', '').strip()
    phone       = data.get('phone', '').strip()

    if not all([national_id, role, name, email]):
        return jsonify({'error': 'Required fields missing'}), 400

    if role == 'admin':
        user = Admin.query.filter_by(national_id=national_id).first()
    elif role == 'specialist':
        user = Specialist.query.filter_by(national_id=national_id).first()
    elif role in ('patient', 'Patient'):
        user = Patient.query.filter_by(national_id=national_id).first()
    else:
        user = RegisteredUser.query.filter_by(national_id=national_id).first()

    if not user:
        return jsonify({'error': 'User not found'}), 404

    # Patients don't have email in our schema; only validate uniqueness for roles that store it
    if role in ('patient', 'Patient'):
        user.name = name
        db.session.commit()
        print(f"✅ Patient profile updated: {name} ({national_id})")
        return jsonify({'message': 'Profile updated successfully', 'user': user.to_dict()}), 200

    # Check email not already taken by someone else
    if role == 'admin':
        existing = Admin.query.filter_by(email=email).first()
    elif role == 'specialist':
        existing = Specialist.query.filter_by(email=email).first()
    else:
        existing = RegisteredUser.query.filter_by(email=email).first()

    if existing and existing.national_id != national_id:
        return jsonify({'error': 'Email already in use'}), 409

    user.name  = name
    user.email = email
    if hasattr(user, 'phone') and phone:
        user.phone = phone
    if hasattr(user, 'phone_num') and phone:
        user.phone_num = phone

    db.session.commit()
    print(f"✅ Profile updated: {name} ({national_id})")
    return jsonify({'message': 'Profile updated successfully', 'user': user.to_dict()}), 200


# ══════════════════════════════════════════════════════════
#   PATIENT SETTINGS (RegisteredUser)
# ══════════════════════════════════════════════════════════

def _get_or_create_patient_settings(national_id: str) -> PatientSettings:
    s = PatientSettings.query.filter_by(user_national_id=national_id).first()
    if s:
        return s
    s = PatientSettings(user_national_id=national_id)
    db.session.add(s)
    db.session.commit()
    return s


@app.route('/patient/settings', methods=['GET'])
def get_patient_settings():
    national_id = str(request.args.get('national_id', '') or '').strip()
    if not national_id:
        return jsonify({'error': 'national_id is required'}), 400
    user = RegisteredUser.query.filter_by(national_id=national_id).first()
    patient = Patient.query.filter_by(national_id=national_id).first()
    if not user and not patient:
        return jsonify({'error': 'User not found'}), 404
    s = _get_or_create_patient_settings(national_id)
    return jsonify(s.to_dict()), 200


@app.route('/patient/settings', methods=['PUT'])
def put_patient_settings():
    data = request.get_json(silent=True) or {}
    national_id = str(data.get('national_id', '') or '').strip()
    if not national_id:
        return jsonify({'error': 'national_id is required'}), 400
    user = RegisteredUser.query.filter_by(national_id=national_id).first()
    patient = Patient.query.filter_by(national_id=national_id).first()
    if not user and not patient:
        return jsonify({'error': 'User not found'}), 404

    s = _get_or_create_patient_settings(national_id)

    def b(key: str, default: bool) -> bool:
        v = data.get(key, default)
        return bool(v)

    try:
        s.notify_hunger = b('notify_hunger', s.notify_hunger)
        s.notify_thirst = b('notify_thirst', s.notify_thirst)
        s.notify_alarm = b('notify_alarm', s.notify_alarm)
        s.notify_bathroom = b('notify_bathroom', s.notify_bathroom)
        s.notify_medicine = b('notify_medicine', s.notify_medicine)

        mc = data.get('min_confidence', None)
        if mc is not None:
            mc_f = float(mc)
            if mc_f < 0 or mc_f > 1:
                return jsonify({'error': 'min_confidence must be between 0 and 1'}), 400
            s.min_confidence = Decimal(str(round(mc_f, 4)))

        if 'text_size' in data:
            ts = str(data.get('text_size') or '').strip().lower()
            if ts not in ('normal', 'large'):
                return jsonify({'error': 'text_size must be normal or large'}), 400
            s.text_size = ts

        if 'high_contrast' in data:
            s.high_contrast = bool(data.get('high_contrast'))

        dr = data.get('data_retention_days', None)
        if dr is not None:
            dr_i = int(dr)
            if dr_i < 7 or dr_i > 3650:
                return jsonify({'error': 'data_retention_days must be between 7 and 3650'}), 400
            s.data_retention_days = dr_i

        if 'preferred_device' in data:
            s.preferred_device = str(data.get('preferred_device') or 'EPOC X').strip()[:50]

        if 'recorded_data_usage_allowed' in data:
            s.recorded_data_usage_allowed = bool(data.get('recorded_data_usage_allowed'))

        s.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'message': 'Settings saved', 'settings': s.to_dict()}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/patient/change-password', methods=['PUT'])
def patient_change_password():
    data = request.get_json(silent=True) or {}
    national_id = str(data.get('national_id', '') or '').strip()
    current_password = str(data.get('current_password', '') or '')
    new_password = str(data.get('new_password', '') or '')
    if not national_id or not current_password or not new_password:
        return jsonify({'error': 'national_id, current_password, and new_password are required'}), 400

    user       = RegisteredUser.query.filter_by(national_id=national_id).first()
    specialist = Specialist.query.filter_by(national_id=national_id).first()
    admin      = Admin.query.filter_by(national_id=national_id).first()

    target = user  or specialist or admin
    if not target:
        return jsonify({'error': 'User not found'}), 404
    if not check_password_hash(target.password, current_password):
        return jsonify({'error': 'Current password is incorrect'}), 403
    if len(new_password) < 8:
        return jsonify({'error': 'New password must be at least 8 characters'}), 400

    target.password = generate_password_hash(new_password)
    db.session.commit()
    return jsonify({'message': 'Password updated'}), 200


@app.route('/patient/sessions/export', methods=['GET'])
def patient_export_sessions_csv():
    patient_national_id = str(request.args.get('patient_national_id', '') or '').strip()
    if not patient_national_id:
        return jsonify({'error': 'patient_national_id is required'}), 400

    sessions = EEGSession.query.filter_by(patient_national_id=patient_national_id).order_by(EEGSession.session_id.desc()).limit(1000).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['session_id','patient_national_id','specialist_national_id','model_id','start_time','end_time','detected_word','confidence_level','device','channels','session_status'])
    for s in sessions:
        writer.writerow([
            s.session_id,
            s.patient_national_id,
            s.specialist_national_id or '',
            s.model_id or '',
            s.start_time.isoformat() if s.start_time else '',
            s.end_time.isoformat() if s.end_time else '',
            s.detected_word,
            float(s.confidence_level) if s.confidence_level is not None else '',
            s.device or '',
            s.channels or 14,
            s.session_status,
        ])

    csv_data = output.getvalue()
    return Response(
        csv_data,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=\"sessions_{patient_national_id}.csv\"'},
    )


# ══════════════════════════════════════════════════════════
#   NOTIFICATIONS (Recipient bell)
# ══════════════════════════════════════════════════════════

def _notification_word_enabled(settings: PatientSettings, detected_word: str) -> bool:
    """
    Map detected word (Arabic) to the recipient's notification toggles (Alerts & safety).
    Used with min_confidence in /notifications/event to decide bell entries.
    """
    w = (detected_word or '').strip()
    if not w:
        return False

    # Arabic vocab used in demo:
    # 'حمام' bathroom, 'إنذار'/'انذار' alarm, 'عطش' thirst, 'جوع' hunger, 'دواء' medicine
    if w == 'جوع':
        return bool(settings.notify_hunger)
    if w == 'عطش':
        return bool(settings.notify_thirst)
    if w in ('انذار', 'إنذار'):
        return bool(settings.notify_alarm)
    if w == 'حمام':
        return bool(settings.notify_bathroom)
    if w == 'دواء':
        return bool(settings.notify_medicine)

    # For 'نعم' / 'لا' or unknown words: never notify by default.
    return False


def _parse_event_time(v) -> datetime:
    if isinstance(v, datetime):
        return v
    s = str(v or '').strip()
    if not s:
        return datetime.utcnow()
    try:
        return datetime.fromisoformat(s.replace('Z', '+00:00'))
    except Exception:
        return datetime.utcnow()


@app.route('/notifications/event', methods=['POST'])
def notifications_event():
    data = request.get_json(silent=True) or {}
    patient_national_id = str(data.get('patient_national_id', '') or '').strip()
    detected_word = str(data.get('detected_word', '') or '').strip()
    confidence_raw = data.get('confidence', None)
    event_time = _parse_event_time(data.get('event_time', None))

    if not patient_national_id:
        return jsonify({'error': 'patient_national_id is required'}), 400
    if not detected_word:
        return jsonify({'error': 'detected_word is required'}), 400

    try:
        conf = float(confidence_raw) if confidence_raw is not None else None
    except Exception:
        conf = None

    # Ensure settings row exists and apply policy.
    settings = _get_or_create_patient_settings(patient_national_id)
    min_conf = float(settings.min_confidence) if settings.min_confidence is not None else 0.0

    # Always record raw event (history / future rules).
    try:
        db.session.add(NotificationEvent(
            patient_national_id=patient_national_id,
            detected_word=detected_word,
            confidence=(Decimal(str(conf)) if conf is not None else None),
            event_time=event_time,
        ))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to record notification event: {e}'}), 500

    created = False
    notification_row = None

    # Bell notifications: (1) word toggle on, (2) confidence >= minimum. No extra consecutive gate.
    if _notification_word_enabled(settings, detected_word) and (conf is not None and conf >= min_conf):
        # Basic de-dupe: don't create two notifications for the same word within the last second
        recent = (
            Notification.query
            .filter_by(patient_national_id=patient_national_id, detected_word=detected_word, seen=False)
            .order_by(Notification.event_time.desc(), Notification.notification_id.desc())
            .first()
        )
        if recent and recent.event_time and abs((event_time - recent.event_time).total_seconds()) < 1.0:
            created = False
            notification_row = None
        else:
            try:
                row = Notification(
                    patient_national_id=patient_national_id,
                    detected_word=detected_word,
                    confidence=(Decimal(str(conf)) if conf is not None else None),
                    event_time=event_time,
                    seen=False,
                )
                db.session.add(row)
                db.session.commit()
                created = True
                notification_row = row
            except Exception as e:
                db.session.rollback()
                return jsonify({'error': f'Failed to create notification: {e}'}), 500

    unseen_count = (
        Notification.query
        .filter_by(patient_national_id=patient_national_id, seen=False)
        .count()
    )

    return jsonify({
        'created': bool(created),
        'notification': (notification_row.to_dict() if notification_row else None),
        'unseen_count': int(unseen_count),
    }), 200


@app.route('/notifications', methods=['GET'])
def list_notifications():
    national_id = str(request.args.get('national_id', '') or '').strip()
    if not national_id:
        return jsonify({'error': 'national_id is required'}), 400
    limit = request.args.get('limit', default=20, type=int)
    offset = request.args.get('offset', default=0, type=int)
    limit = max(1, min(int(limit), 100))
    offset = max(0, int(offset))

    q = Notification.query.filter_by(patient_national_id=national_id).order_by(Notification.event_time.desc(), Notification.notification_id.desc())
    rows = q.offset(offset).limit(limit).all()
    unseen_count = Notification.query.filter_by(patient_national_id=national_id, seen=False).count()
    return jsonify({
        'items': [r.to_dict() for r in rows],
        'unseen_count': int(unseen_count),
    }), 200


@app.route('/notifications/<int:notification_id>/seen', methods=['PUT'])
def mark_notification_seen(notification_id: int):
    data = request.get_json(silent=True) or {}
    national_id = str(data.get('national_id', '') or '').strip()
    if not national_id:
        return jsonify({'error': 'national_id is required'}), 400

    row = Notification.query.filter_by(notification_id=notification_id, patient_national_id=national_id).first()
    if not row:
        return jsonify({'error': 'Notification not found'}), 404

    row.seen = True
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to mark seen: {e}'}), 500

    unseen_count = Notification.query.filter_by(patient_national_id=national_id, seen=False).count()
    return jsonify({'message': 'ok', 'unseen_count': int(unseen_count)}), 200


@app.route('/notifications/seen-all', methods=['PUT'])
def mark_all_notifications_seen():
    data = request.get_json(silent=True) or {}
    national_id = str(data.get('national_id', '') or '').strip()
    if not national_id:
        return jsonify({'error': 'national_id is required'}), 400
    try:
        Notification.query.filter_by(patient_national_id=national_id, seen=False).update(
            {'seen': True},
            synchronize_session=False,
        )
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to mark all seen: {e}'}), 500
    return jsonify({'message': 'ok', 'unseen_count': 0}), 200


@app.route('/specialist/patient-notifications', methods=['GET'])
def specialist_patient_notifications():
    """Recent notifications for all patients assigned to this specialist (clinical needs / bell aggregate)."""
    specialist_id = str(request.args.get('specialist_id', '') or '').strip()
    if not specialist_id:
        return jsonify({'error': 'specialist_id is required'}), 400
    limit = request.args.get('limit', default=30, type=int)
    limit = max(1, min(int(limit), 100))

    try:
        sid_int = int(specialist_id)
    except Exception:
        sid_int = None

    spec = Specialist.query.filter(
        or_(Specialist.national_id == specialist_id, Specialist.national_id == sid_int)
    ).first()
    if not spec:
        return jsonify({'error': 'Specialist not found'}), 404
    spec_key = str(spec.national_id)

    pat_conds = [Patient.specialist_national_id == spec_key]
    if sid_int is not None:
        pat_conds.append(Patient.specialist_national_id == sid_int)
    patients = Patient.query.filter(or_(*pat_conds)).all()
    if not patients:
        return jsonify({'items': [], 'unseen_count': 0}), 200

    name_by: dict[str, str] = {}
    pid_keys: list[str] = []
    for p in patients:
        s = str(p.national_id)
        name_by[s] = p.name or 'Patient'
        pid_keys.append(s)

    notif_conds = []
    for k in pid_keys:
        notif_conds.append(Notification.patient_national_id == k)
        try:
            notif_conds.append(Notification.patient_national_id == int(k))
        except Exception:
            pass

    rows = (
        Notification.query.filter(or_(*notif_conds))
        .order_by(Notification.event_time.desc(), Notification.notification_id.desc())
        .limit(limit)
        .all()
    )

    unseen_total = (
        Notification.query.filter(or_(*notif_conds), Notification.seen.is_(False)).count()
        if notif_conds
        else 0
    )

    out = []
    for r in rows:
        pk = str(r.patient_national_id)
        d = r.to_dict()
        pname = name_by.get(pk)
        if not pname and pk.isdigit():
            pname = name_by.get(str(int(pk)))
        d['patient_name'] = pname or 'Patient'
        out.append(d)

    return jsonify({'items': out, 'unseen_count': int(unseen_total)}), 200


@app.route('/specialist/patient-notifications/seen-all', methods=['PUT'])
def specialist_mark_all_patient_notifications_seen():
    data = request.get_json(silent=True) or {}
    specialist_id = str(data.get('specialist_id', '') or '').strip()
    if not specialist_id:
        return jsonify({'error': 'specialist_id is required'}), 400
    try:
        sid_int = int(specialist_id)
    except Exception:
        sid_int = None
    spec = Specialist.query.filter(
        or_(Specialist.national_id == specialist_id, Specialist.national_id == sid_int)
    ).first()
    if not spec:
        return jsonify({'error': 'Specialist not found'}), 404
    spec_key = str(spec.national_id)
    pat_conds = [Patient.specialist_national_id == spec_key]
    if sid_int is not None:
        pat_conds.append(Patient.specialist_national_id == sid_int)
    patients = Patient.query.filter(or_(*pat_conds)).all()
    try:
        for p in patients:
            pid = p.national_id
            Notification.query.filter(
                or_(Notification.patient_national_id == str(pid), Notification.patient_national_id == pid),
                Notification.seen.is_(False),
            ).update({'seen': True}, synchronize_session=False)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to mark all seen: {e}'}), 500
    return jsonify({'message': 'ok', 'unseen_count': 0}), 200


# ══════════════════════════════════════════════════════════
# START SERVER
# ══════════════════════════════════════════════════════════


if __name__ == '__main__':

    with app.app_context():
        db.create_all()  #create the database
        ensure_registered_user_verification_columns()
        ensure_admin_specialist_gender_columns()
        ensure_admin_specialist_verification_columns()
        ensure_patient_room_numbers()
        ensure_patient_password_column()
        ensure_patient_settings_recorded_data_usage_column()
        ensure_specialist_3030303030_demo_patient_and_sessions()
        ensure_demo_patient_notifications_for_specialist()
        print("[OK] Database tables ready")
        print("\nAPI listening on http://127.0.0.1:5000 (and http://localhost:5000)\n")

    app.run(debug=True, host='0.0.0.0', port=5000)